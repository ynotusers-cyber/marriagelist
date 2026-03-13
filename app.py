import streamlit as st
import sqlite3
import pandas as pd
import os
import openpyxl

# Paths
BASE_DIR = os.path.dirname(__file__)
DB_PATH = os.path.join(BASE_DIR, 'seeru.db')
EXCEL_PATH = os.path.join(BASE_DIR, 'SSYADAV-6326.xlsx')

# DB helpers
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS functions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            date TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            place TEXT,
            mobile TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS received (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id INTEGER NOT NULL,
            function_id INTEGER NOT NULL,
            amount REAL NOT NULL,
            mode TEXT DEFAULT 'CASH',
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(contact_id) REFERENCES contacts(id),
            FOREIGN KEY(function_id) REFERENCES functions(id)
        );
        CREATE TABLE IF NOT EXISTS given (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contact_id INTEGER NOT NULL,
            event_name TEXT NOT NULL,
            event_date TEXT,
            amount REAL NOT NULL,
            mode TEXT DEFAULT 'CASH',
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(contact_id) REFERENCES contacts(id)
        );
    ''')
    conn.commit()
    conn.close()

# Excel import
def import_excel():
    if not os.path.exists(EXCEL_PATH):
        st.warning("Excel file not found!")
        return 0
    conn = get_db()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    func_title = ws.cell(row=1, column=1).value or "SS YADAV FAMILY FUNCTION - MARCH 2026"
    cur = conn.execute('INSERT INTO functions (title,date) VALUES (?,?)', (func_title,'2026-03-06'))
    func_id = cur.lastrowid
    imported = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        slno, place, name, mobile, amount, mode = row[:6]
        if not name or not amount: continue
        cur = conn.execute(
            'INSERT INTO contacts (name,place,mobile) VALUES (?,?,?)',
            (str(name).strip(), str(place).strip() if place else '', str(int(mobile)) if isinstance(mobile,float) else (str(mobile) if mobile else ''))
        )
        contact_id = cur.lastrowid
        conn.execute(
            'INSERT INTO received (contact_id,function_id,amount,mode) VALUES (?,?,?,?)',
            (contact_id, func_id, float(amount), str(mode).strip() if mode else 'CASH')
        )
        imported += 1
    conn.commit()
    conn.close()
    return imported

# Fetch stats
def fetch_stats():
    conn = get_db()
    total_received = conn.execute('SELECT COALESCE(SUM(amount),0) FROM received').fetchone()[0]
    total_given = conn.execute('SELECT COALESCE(SUM(amount),0) FROM given').fetchone()[0]
    total_contacts = conn.execute('SELECT COUNT(*) FROM contacts').fetchone()[0]
    total_functions = conn.execute('SELECT COUNT(*) FROM functions').fetchone()[0]
    conn.close()
    return {
        'Total Received': total_received,
        'Total Given': total_given,
        'Balance': total_received - total_given,
        'Total Contacts': total_contacts,
        'Total Functions': total_functions
    }

# Fetch contacts
def fetch_contacts(search=''):
    conn = get_db()
    q = f'%{search.lower()}%'
    if search:
        rows = conn.execute('''
            SELECT c.id, c.name, c.place, c.mobile,
            COALESCE(SUM(r.amount),0) as total_received,
            COALESCE((SELECT SUM(g.amount) FROM given g WHERE g.contact_id=c.id),0) as total_given
            FROM contacts c
            LEFT JOIN received r ON r.contact_id=c.id
            WHERE LOWER(c.name) LIKE ? OR LOWER(c.place) LIKE ? OR c.mobile LIKE ?
            GROUP BY c.id
        ''',(q,q,q)).fetchall()
    else:
        rows = conn.execute('''
            SELECT c.id, c.name, c.place, c.mobile,
            COALESCE(SUM(r.amount),0) as total_received,
            COALESCE((SELECT SUM(g.amount) FROM given g WHERE g.contact_id=c.id),0) as total_given
            FROM contacts c
            LEFT JOIN received r ON r.contact_id=c.id
            GROUP BY c.id
        ''').fetchall()
    conn.close()
    return pd.DataFrame([dict(r) for r in rows])

# Fetch contact details
def fetch_contact_details(cid):
    conn = get_db()
    contact = conn.execute('SELECT * FROM contacts WHERE id=?',(cid,)).fetchone()
    received = conn.execute('''
        SELECT r.amount, r.mode, r.notes, f.title AS event_name, r.created_at
        FROM received r
        JOIN functions f ON r.function_id=f.id
        WHERE r.contact_id=?
    ''',(cid,)).fetchall()
    given = conn.execute('''
        SELECT amount, mode, event_name, event_date, notes, created_at
        FROM given WHERE contact_id=?
    ''',(cid,)).fetchall()
    conn.close()
    return dict(contact), [dict(r) for r in received], [dict(g) for g in given]

# ── Streamlit UI ──
st.set_page_config(page_title="Seeru Book", layout="wide")
st.title("💰 Seeru Book Dashboard")
init_db()

menu = st.sidebar.radio("Menu",["Dashboard","Contacts","Add Contact","Import Excel"])

if menu=="Dashboard":
    stats = fetch_stats()
    col1,col2,col3=st.columns(3)
    col1.metric("Total Received",stats['Total Received'])
    col2.metric("Total Given",stats['Total Given'])
    col3.metric("Balance",stats['Balance'])
    st.write(f"Total Contacts: {stats['Total Contacts']} | Total Functions: {stats['Total Functions']}")

elif menu=="Contacts":
    search = st.text_input("Search by name, place, or mobile")
    df = fetch_contacts(search)
    st.dataframe(df,use_container_width=True)
    if st.checkbox("Show contact details"):
        cid = st.number_input("Enter Contact ID",min_value=1,step=1)
        if cid:
            contact,received,given = fetch_contact_details(cid)
            st.subheader("Contact Info")
            st.json(contact)
            st.subheader("Received Gifts")
            st.dataframe(pd.DataFrame(received))
            st.subheader("Given Gifts")
            st.dataframe(pd.DataFrame(given))

elif menu=="Add Contact":
    with st.form("add_contact_form"):
        name = st.text_input("Name")
        place = st.text_input("Place")
        mobile = st.text_input("Mobile")
        submitted = st.form_submit_button("Add Contact")
        if submitted:
            if not name.strip():
                st.error("Name required")
            else:
                conn = get_db()
                conn.execute('INSERT INTO contacts(name,place,mobile) VALUES(?,?,?)',(name.strip(),place.strip(),mobile.strip()))
                conn.commit()
                conn.close()
                st.success(f"Contact '{name}' added!")

elif menu=="Import Excel":
    st.info("Import contacts from Excel")
    if st.button("Import Now"):
        imported = import_excel()
        st.success(f"Imported {imported} contacts from Excel")